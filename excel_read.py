import csv
import logging as log
import re
from openpyxl import Workbook

log.basicConfig(level=log.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

csv_file_path = r"C:\Users\admin\OneDrive\바탕 화면\dup_clikey.csv"
excel_file_path = r"C:\Users\admin\OneDrive\바탕 화면\dup_clikey_result.xlsx"

# 정규식 패턴
regex = r'"cliKey":"(.*?)"'

# 엑셀파일 쓰기
write_wb = Workbook()
write_ws = write_wb.active
write_ws.title = 'cliKey'

with open(csv_file_path, newline='', encoding='utf-8') as csvfile:
    csvreader = csv.reader(csvfile)
    row_num = 1
    for i, row in enumerate(csvreader):
        match =  re.search(regex, row[0])
        if match:
            cliKey = match.group(1)  # 매칭된 tmpltCode 값
            write_ws.cell(row=row_num, column=1, value='cliKey 값')
            write_ws.cell(row=row_num, column=2, value=cliKey)
            row_num += 1

write_wb.save(excel_file_path)
log.info(f'{excel_file_path} 파일이 저장되었습니다.')
