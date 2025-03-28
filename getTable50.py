from openpyxl import Workbook
from src.client.schema_info_provider import SchemaInfoProvider

def create_excel_with_table_data(schema, output_file):
    provider = SchemaInfoProvider()
    tables = provider.get_tables(schema)

    wb = Workbook()
    for table in tables:
        table_name = table['table_name']
        data = provider.get_data(schema, table_name, 50)

        # if not data:
        #     continue  # Skip if no data is returned

        ws = wb.create_sheet(title=table_name)

        # Write headers
        headers = list(data[0].keys())
        ws.append(headers)

        # Write data rows
        for row in data:
            encoded_row = [str(value).encode('utf-8', errors='ignore').decode('utf-8') for value in row.values()]
            ws.append(encoded_row)

    # Remove the default sheet created by openpyxl
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    wb.save(output_file)

if __name__ == "__main__":
    create_excel_with_table_data('cm_push', 'download.xlsx')