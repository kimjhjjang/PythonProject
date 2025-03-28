import logging as log
import sys
import os
import smtplib

sys.path.append(r"C:\Users\admin\PycharmProjects\PythonProject\.venv\Lib\site-packages")
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

# path setting
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), 'src')))
from src.client.schema_info_provider import SchemaInfoProvider

# Configure logging
log.basicConfig(level=log.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Create an instance of SchemaInfoProvider
provider = SchemaInfoProvider()

# Call the get_daily_data function
daily_data = provider.get_daily_data()

# Log the results
log.info(f'Daily Data: {daily_data}')

# Get the current date in YYYYmmdd format
current_date = datetime.now().strftime('%Y%m%d')
current_date_mmdd = datetime.now().strftime('%m.%d')


# Define the Excel file path with the current date
excel_file_path = rf"C:\Users\admin\OneDrive\바탕 화면\고객사발송한도별사용금액현황_{current_date}.xlsx"

wb = Workbook()
ws = wb.active
ws.title = 'Daily Data'

# Define styles
header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
center_alignment = Alignment(horizontal='center', vertical='center')

# Write the headers
headers = list(daily_data[0].keys())
ws.append(headers)

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center_alignment
    ws.column_dimensions[get_column_letter(col_num)].auto_size = True

# Write the data
for row_num, data in enumerate(daily_data, 2):
    for col_num, value in enumerate(data.values(), 1):
        cell = ws.cell(row=row_num, column=col_num, value=value)
        cell.border = thin_border

# Adjust column widths
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 10)
    ws.column_dimensions[column].width = adjusted_width

# Save the workbook to the specified file path
wb.save(excel_file_path)

log.info(f'Excel file saved at {excel_file_path}')

def send_email(sender_email, receiver_email, subject, body, attachment_path):
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = receiver_email
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))

    attachment = open(attachment_path, 'rb')
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(attachment.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename= {os.path.basename(attachment_path)}')
    msg.attach(part)

    try:
        # server = smtplib.SMTP('smtp.gmail.com', 587) 나의 구글 SMTP
        server = smtplib.SMTP('webmail.lguplus.co.kr', 25)
        server.starttls()
        # server.login(sender_email, 'ywiq gdin qyjr nkoz')  나의 구글 ID
        server.login(sender_email, 'dlfekdqor12%')
        text = msg.as_string()
        server.sendmail(sender_email, receiver_email, text)
        server.quit()
        log.info('Email sent successfully')
    except Exception as e:
        log.error(f'Error sending email: {e}')

# Email configuration
#sender_email = 'kimjhjjang@gmail.com'
sender_email = 'immaster1@lgupluspartners.co.kr'
receiver_email = 'kimjhjjang@naver.com'
subject = f'[메시지허브] 고객사 발송한도 별 사용금액 현황({current_date_mmdd})'
# Check for customers with percent greater than 0
customers_above_zero = [data for data in daily_data if data['percent'] > 0]

if customers_above_zero:
    body = "사용률 0이 넘는 고객사는 다음과 같습니다.\n\n"
    headers = list(customers_above_zero[0].keys())
    body += "\t".join(headers) + "\n"
    for customer in customers_above_zero:
        body += "\t".join(str(customer[header]) for header in headers) + "\n"
else:
    body = "사용률이 0보다 큰 고객사는 없습니다."

log.info(f'Email body: {body}')

# Send the email
send_email(sender_email, receiver_email, subject, body, excel_file_path)